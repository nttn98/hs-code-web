"""
Robust Excel -> JSON mapper (outline-level from column A).

Usage examples:
    # preview mapping only
    python excel_colletter_map.py -i data/hscode.xlsx -hr 2 --preview-only

    # full conversion (writes outputs)
    python excel_colletter_map.py -i data/hscode.xlsx -hr 2

Outputs:
  - per-row JSON: default data/hscode_by_colletter.json
  - grouped JSON:  data/hscode_grouped.json (root + children include mo_ta_vn/mo_ta_en)
  - compact searchable nodes: data/hscode_grouped_tree.json
"""
from __future__ import annotations
import argparse
import json
import re
import unicodedata
from typing import Dict, List, Optional, Tuple, Any
from openpyxl import load_workbook

# ---------- fallback column-letter -> schema key (used only if header detection fails) ----------
COL_MAP: Dict[str, str] = {
    "A": "v",
    "B": "ma_hang",
    "C": "mo_ta_hang",
    "D": "mo_ta_hang_en",
    "E": "don_vi_tinh",
    "F": "unit_of_quantity",
    "G": "nk_tt",
    "H": "nk_tt_van_ban",
    "I": "nk_tt_ngay_hieu_luc",
    "J": "nk_uu_dai",
    "K": "nk_uu_dai_van_ban",
    "L": "nk_uu_dai_ngay_hieu_luc",
    "M": "vat",
    "N": "vat_van_ban",
    "O": "vat_ngay_hieu_luc",
    "P": "acfta",
    "Q": "acfta_van_ban",
    "R": "acfta_ngay_hieu_luc",
    "S": "atiga",
    "T": "atiga_van_ban",
    "U": "atiga_ngay_hieu_luc",
    "V": "ajcep",
    "W": "ajcep_van_ban",
    "X": "ajcep_ngay_hieu_luc",
    "Y": "vjepa",
    "Z": "vjepa_van_ban",
    "AA": "vjepa_ngay_hieu_luc",
    "AB": "akfta",
    "AC": "akfta_van_ban",
    "AD": "akfta_ngay_hieu_luc",
    "AE": "aanzfta",
    "AF": "aanzfta_van_ban",
    "AG": "aanzfta_ngay_hieu_luc",
    "AH": "aifta",
    "AI": "aifta_van_ban",
    "AJ": "aifta_ngay_hieu_luc",
    "AK": "vkfta",
    "AL": "vkfta_van_ban",
    "AM": "vkfta_ngay_hieu_luc",
    "AN": "vcfta",
    "AO": "vcfta_van_ban",
    "AP": "vcfta_ngay_hieu_luc",
    "AQ": "vn_eaeu",
    "AR": "vn_eaeu_van_ban",
    "AS": "vn_eaeu_ngay_hieu_luc",
    "AT": "cptpp",
    "AU": "cptpp_van_ban",
    "AV": "cptpp_ngay_hieu_luc",
    "AW": "ahkfta",
    "AX": "ahkfta_van_ban",
    "AY": "ahkfta_ngay_hieu_luc",
    "AZ": "vncu",
    "BA": "vncu_van_ban",
    "BB": "vncu_ngay_hieu_luc",
    "BC": "evfta",
    "BD": "evfta_van_ban",
    "BE": "evfta_ngay_hieu_luc",
    "BF": "ukvfta",
    "BG": "ukvfta_van_ban",
    "BH": "ukvfta_ngay_hieu_luc",
    "BI": "vn_lao_raw",
    "BJ": "vn_lao_raw_van_ban",
    "BK": "vn_lao_raw_ngay_hieu_luc",
    "BL": "vifta",
    "BM": "vifta_van_ban",
    "BN": "vifta_ngay_hieu_luc",
    "BO": "rcept",
    "BP": "rcept_van_ban",
    "BQ": "rcept_ngay_hieu_luc",
    "BR": "ttdb",
    "BS": "ttdb_van_ban",
    "BT": "ttdb_ngay_hieu_luc",
    "BU": "xk",
    "BV": "xk_van_ban",
    "BW": "xk_ngay_hieu_luc",
    "BX": "xk_cptpp",
    "BY": "xk_cptpp_van_ban",
    "BZ": "xk_cptpp_ngay_hieu_luc",
    "CA": "xk_ev",
    "CB": "xk_ev_van_ban",
    "CC": "xk_ev_ngay_hieu_luc",
    "CD": "xk_ukv",
    "CE": "xk_ukv_van_ban",
    "CF": "xk_ukv_ngay_hieu_luc",
    "CG": "thue_bvmt",
    "CH": "thue_bvmt_van_ban",
    "CI": "thue_bvmt_ngay_hieu_luc",
    "CJ": "chinh_sach_ma_hs",
    "CK": "giam_vat",
    "CL": "chi_tiet_giam_vat",
}
# ------------------------------------------------------------------------------------------

_COLLETTER_RE = re.compile(r"^[A-Z]{1,3}$")

def index_to_colletter(idx: int) -> str:
    s = ""
    n = idx + 1
    while n > 0:
        n, rem = divmod(n - 1, 26)
        s = chr(rem + ord("A")) + s
    return s

def normalize_text(s: Optional[Any]) -> str:
    if s is None:
        return ""
    s = str(s)
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return s.strip()

def detect_level_from_v(text: Optional[Any]) -> int:
    """Fallback: count leading '-' in v_raw (legacy)."""
    if text is None:
        return 0
    s = normalize_text(text).lstrip()
    if s == "":
        return 0
    s2 = s.replace("–", "-").replace("—", "-")
    m = re.match(r"^(-\s*)+", s2)
    if not m:
        return 0
    return m.group(0).count("-")

def clean_code(s: Optional[Any]) -> Optional[str]:
    if s is None:
        return None
    s2 = re.sub(r"\D", "", str(s))
    return s2 or None

def read_header_block(path: str, header_rows: int) -> Tuple[List[str], int, Optional[int], Optional[List[Optional[str]]]]:
    wb = load_workbook(filename=path, read_only=True, data_only=True)
    ws = wb.active
    header_vals = []
    max_col = 0
    for r in ws.iter_rows(min_row=1, max_row=header_rows, values_only=True):
        row = [("" if v is None else str(v).strip()) for v in r]
        max_col = max(max_col, len(row))
        header_vals.append(row)
    wb.close()
    # pad
    for i in range(len(header_vals)):
        if len(header_vals[i]) < max_col:
            header_vals[i] += [""] * (max_col - len(header_vals[i]))
    eff_headers = []
    for c in range(max_col):
        parts = []
        for r in range(header_rows):
            v = header_vals[r][c].strip()
            if v != "":
                parts.append(v)
        eff_headers.append(" | ".join(parts))
    # detect explicit letters row
    explicit_row_idx = None
    explicit_letters = None
    for r in range(header_rows):
        row = header_vals[r]
        nonempty = sum(1 for v in row if v != "")
        if nonempty == 0:
            continue
        letter_like = sum(1 for v in row if v and _COLLETTER_RE.match(str(v).strip().upper()))
        pct = letter_like / nonempty
        if pct >= 0.5:
            explicit_row_idx = r
            explicit_letters = []
            for v in row:
                if v is None:
                    explicit_letters.append(None)
                else:
                    vv = str(v).strip().upper()
                    explicit_letters.append(vv if _COLLETTER_RE.match(vv) else None)
            break
    return eff_headers, max_col, explicit_row_idx, explicit_letters

def build_idx_to_letter(max_col: int, explicit_letters: Optional[List[Optional[str]]]) -> Dict[int, str]:
    idx_to_letter = {}
    for i in range(max_col):
        if explicit_letters and i < len(explicit_letters) and explicit_letters[i]:
            idx_to_letter[i] = explicit_letters[i]
        else:
            idx_to_letter[i] = index_to_colletter(i)
    return idx_to_letter

def map_index_to_schema(eff_headers: List[str], idx_to_letter: Dict[int, str]) -> Dict[int, str]:
    header_lower = [normalize_text(h).lower() for h in eff_headers]
    header_name_map = {
        "mã hàng": "ma_hang",
        "ma hang": "ma_hang",
        "mô tả": "mo_ta_hang",
        "mô tả hàng": "mo_ta_hang",
        "tiếng anh": "mo_ta_hang_en",
        "english": "mo_ta_hang_en",
        "đơn vị": "don_vi_tinh",
        "unit of quantity": "unit_of_quantity",
        "vat": "vat",
        "nk tt": "nk_tt",
        "nk ưu": "nk_uu_dai",
        "tt đb": "ttdb",
        "thuế bvmt": "thue_bvmt",
        "xk": "xk",
    }
    idx_map = {}
    for idx, hdr in enumerate(header_lower):
        if hdr and hdr != "":
            for key, sk in header_name_map.items():
                if key in hdr:
                    idx_map[idx] = sk
                    break
    # fallback by explicit letters -> COL_MAP
    for idx in range(len(idx_to_letter)):
        if idx in idx_map:
            continue
        letter = idx_to_letter.get(idx)
        if letter and letter in COL_MAP:
            idx_map[idx] = COL_MAP[letter]
    return idx_map

def best_desc_indices(eff_headers: List[str]) -> Tuple[Optional[int], Optional[int]]:
    # detect VN and EN desc columns, robust heuristics
    desc_vn_idx = None
    desc_en_idx = None
    for i, hdr in enumerate(eff_headers):
        if not hdr:
            continue
        low = normalize_text(hdr).lower()
        if desc_vn_idx is None and ("mô tả" in low or "mota" in low or "mo ta" in low):
            # prefer column mentions "tiếng việt" or no "tiếng anh"
            if "tiếng anh" not in low and "english" not in low:
                desc_vn_idx = i
        if desc_en_idx is None and ("tiếng anh" in low or "english" in low or "tiếng anh" in low):
            desc_en_idx = i
    # fallback: typical layout C -> D (0-based: 2,3)
    if desc_vn_idx is None and len(eff_headers) > 2:
        desc_vn_idx = 2
    if desc_en_idx is None and len(eff_headers) > 3:
        # but only set en if different from vn
        if desc_en_idx is None and desc_vn_idx != 3:
            desc_en_idx = 3
    return desc_vn_idx, desc_en_idx

def parse_level_from_first_cell(cell_value: Any) -> int:
    """
    Parse the level number from first column (cột A).
    Accepts int-like or string digits. Fallback to detect_level_from_v if not parseable.
    """
    if cell_value is None:
        return 0
    # try direct int
    try:
        return int(cell_value)
    except Exception:
        pass
    # if it's a string like '0', '1', keep
    s = str(cell_value).strip()
    if re.fullmatch(r"\d+", s):
        return int(s)
    # fallback: detect from v_raw (legacy - dashes)
    return detect_level_from_v(cell_value)

def process(input_path: str, header_rows: int, output_rows: str, output_grouped: str, output_tree: str, preview_only: bool=False):
    print("Reading header block from:", input_path)
    eff_headers, max_col, explicit_row_idx, explicit_letters = read_header_block(input_path, header_rows)
    print("Effective headers (first 40):")
    for i,h in enumerate(eff_headers[:min(80,len(eff_headers))]):
        print(f"{i:03d}: {h[:80]}")
    if explicit_row_idx is not None:
        print(f"Detected explicit column-letter row at header index {explicit_row_idx}")

    idx_to_letter = build_idx_to_letter(max_col, explicit_letters)
    idx_map = map_index_to_schema(eff_headers, idx_to_letter)
    desc_vn_idx, desc_en_idx = best_desc_indices(eff_headers)
    print(f"\nDetected desc_vn_idx={desc_vn_idx}, desc_en_idx={desc_en_idx}")

    print("\nPreview mapping (index -> letter -> header -> schema):")
    for i in range(max_col):
        letter = idx_to_letter.get(i, index_to_colletter(i))
        hdr = eff_headers[i] if i < len(eff_headers) else ""
        mapped = idx_map.get(i)
        extra = ""
        if i == desc_vn_idx: extra += " [DETECTED VN_DESC]"
        if i == desc_en_idx: extra += " [DETECTED EN_DESC]"
        print(f"[{i:03d}] {letter:>4} | '{hdr}' -> {mapped}{extra}")

    if preview_only:
        print("\nPreview-only: exit.")
        return

    wb = load_workbook(filename=input_path, read_only=True, data_only=True)
    ws = wb.active
    rows: List[Dict[str, Any]] = []
    for row_idx, excel_row in enumerate(ws.iter_rows(min_row=header_rows+1, values_only=True), start=header_rows+1):
        cells = list(excel_row) + [None] * max(0, max_col - len(excel_row))
        other_fields = {}
        for idx in range(max_col):
            val = cells[idx] if idx < len(cells) else None
            if val is None: continue
            s = str(val).strip()
            if s == "": continue
            header_name = eff_headers[idx] if idx < len(eff_headers) else index_to_colletter(idx)
            other_fields[header_name] = s

        # find v and ma_hang by mapping or heuristics
        v_raw = None
        ma_raw = None
        for idx, sk in idx_map.items():
            if sk == "v" and idx < len(cells):
                v_raw = cells[idx]
            if sk == "ma_hang" and idx < len(cells):
                ma_raw = cells[idx]

        # heuristic fallbacks
        if v_raw in (None, ""):
            # find column whose header exactly 'V' (case header shows 'V')
            for idx,hdr in enumerate(eff_headers):
                if hdr and hdr.strip().upper() == "V":
                    v_raw = cells[idx] if idx < len(cells) else None
                    break
            # fallback to column A (0)
            if (v_raw is None or str(v_raw).strip()=="") and len(cells) > 0 and cells[0] is not None and str(cells[0]).strip() != "":
                # NOTE: cells[0] is the level column in your file; do not confuse with v_raw.
                # If column A is not actual level but contains v_raw, this preserves behavior.
                v_raw = cells[0]

        if ma_raw in (None, ""):
            for idx,hdr in enumerate(eff_headers):
                low = normalize_text(hdr).lower()
                if "mã hàng" in low or "ma hang" in low or low.startswith("mã") or low.startswith("ma "):
                    ma_raw = cells[idx] if idx < len(cells) else None
                    break
            # fallback try column B
            if (ma_raw is None or str(ma_raw).strip() == "") and len(cells) > 1 and cells[1] is not None:
                cand = str(cells[1]).strip()
                if re.search(r"\d", cand):
                    ma_raw = cand

        mo_ta_vn = ""
        mo_ta_en = ""
        if desc_vn_idx is not None and desc_vn_idx < len(cells) and cells[desc_vn_idx] is not None:
            mo_ta_vn = str(cells[desc_vn_idx]).strip()
        if desc_en_idx is not None and desc_en_idx < len(cells) and cells[desc_en_idx] is not None:
            mo_ta_en = str(cells[desc_en_idx]).strip()

        # parse the explicit level from first cell (column A)
        level_cell_val = cells[0] if len(cells) > 0 else None
        v_level = parse_level_from_first_cell(level_cell_val)

        rec = {
            "row_num": row_idx,
            "other_fields": other_fields,
            "v_raw": v_raw if v_raw is not None else "",
            "v_norm": normalize_text(v_raw) if v_raw is not None else "",
            "v_level": v_level,
            "ma_hang_raw": ma_raw if ma_raw is not None else "",
            "ma_hang_clean": clean_code(ma_raw),
            "mo_ta_vn": mo_ta_vn,
            "mo_ta_en": mo_ta_en,
        }
        rows.append(rec)
    wb.close()

    # infer ma_hang_inferred using previous logic (last seen per level fallback)
    last_seen: Dict[int, str] = {}
    for rec in rows:
        lvl = rec["v_level"] if rec["v_level"] is not None else 0
        if rec["ma_hang_clean"]:
            last_seen[lvl] = rec["ma_hang_clean"]
            # clear deeper levels
            for k in list(last_seen.keys()):
                if k > lvl:
                    del last_seen[k]
            rec["ma_hang_inferred"] = rec["ma_hang_clean"]
        else:
            inferred = None
            if lvl in last_seen:
                inferred = last_seen[lvl]
            else:
                smaller = [L for L in last_seen.keys() if L < lvl]
                if smaller:
                    inferred = last_seen[max(smaller)]
            rec["ma_hang_inferred"] = inferred

    # Build nodes from rows (only rows that have an inferred or explicit code)
    nodes: List[Dict[str, Any]] = []
    for rec in rows:
        code = rec.get("ma_hang_inferred")
        # allow including rows without code? For searchable nodes we skip those
        if not code:
            # still include possible top-level rows that have title but no code? We skip for now.
            continue
        title = re.sub(r'^\-+\s*', '', normalize_text(rec.get("v_raw",""))).strip()
        nodes.append({
            "hs_digits": code,
            "title": title,
            "mo_ta_vn": rec.get("mo_ta_vn", ""),
            "mo_ta_en": rec.get("mo_ta_en", ""),
            "v_level": rec.get("v_level", 0),
            "row_num": rec["row_num"],
            "other_fields": rec["other_fields"],
            "parent": None,
            "children": []
        })

    # Link parent-child using outline stack logic (use v_level from column A)
    # Maintain last node index for each level encountered.
    level_stack: Dict[int, Dict[str, Any]] = {}  # level -> node
    ordered_nodes: List[Dict[str, Any]] = []  # same as nodes, in order
    for node in nodes:
        lvl = node.get("v_level", 0)
        # find parent: nearest smaller level that exists in stack
        parent_node = None
        # search decreasing levels
        for L in range(lvl-1, -1, -1):
            if L in level_stack:
                parent_node = level_stack[L]
                break
        if parent_node:
            node["parent"] = parent_node["hs_digits"]
            if node["hs_digits"] not in parent_node["children"]:
                parent_node["children"].append(node["hs_digits"])
        # push/update this level in stack (overwrite deeper levels)
        level_stack[lvl] = node
        # remove deeper levels > lvl
        for k in list(level_stack.keys()):
            if k > lvl:
                del level_stack[k]
        ordered_nodes.append(node)

    # As a safety-net: if any node still has no parent but its code shares a prefix with
    # a prior node, try linking by nearest prior prefix (fallback)
    for i, node in enumerate(ordered_nodes):
        if node.get("parent") is not None:
            continue
        child_code = node["hs_digits"] or ""
        if not child_code:
            continue
        for j in range(i-1, -1, -1):
            cand = ordered_nodes[j]
            cand_code = cand.get("hs_digits") or ""
            if cand_code and len(cand_code) < len(child_code) and child_code.startswith(cand_code):
                node["parent"] = cand_code
                if node["hs_digits"] not in cand["children"]:
                    cand["children"].append(node["hs_digits"])
                break

    # aggregate nodes per hs and build final search nodes (include VN/EN)
    agg: Dict[str, Dict[str, Any]] = {}
    for n in ordered_nodes:
        k = n["hs_digits"]
        if k not in agg:
            agg[k] = {
                "hs_code": k,
                "titles": [n["title"]] if n["title"] else [],
                "mo_ta_vn": [n["mo_ta_vn"]] if n.get("mo_ta_vn") else [],
                "mo_ta_en": [n["mo_ta_en"]] if n.get("mo_ta_en") else [],
                "parents": set([n["parent"]]) if n["parent"] else set(),
                "children": set(n["children"]),
                "other_samples": [n["other_fields"]],
            }
        else:
            a = agg[k]
            if n["title"] and n["title"] not in a["titles"]:
                a["titles"].append(n["title"])
            if n.get("mo_ta_vn"):
                if n["mo_ta_vn"] not in a["mo_ta_vn"]:
                    a["mo_ta_vn"].append(n["mo_ta_vn"])
            if n.get("mo_ta_en"):
                if n["mo_ta_en"] not in a["mo_ta_en"]:
                    a["mo_ta_en"].append(n["mo_ta_en"])
            if n["parent"]:
                a["parents"].add(n["parent"])
            a["children"].update(n["children"])
            if n["other_fields"]:
                a["other_samples"].append(n["other_fields"])

    final_nodes: List[Dict[str, Any]] = []
    for hs, data in agg.items():
        title = data["titles"][0] if data["titles"] else ""
        parts = [title]
        if data["mo_ta_vn"]:
            parts.append(data["mo_ta_vn"][0])
        if data["mo_ta_en"]:
            parts.append(data["mo_ta_en"][0])
        for sample in data["other_samples"][:2]:
            for hk,hv in sample.items():
                parts.append(f"{hk} {hv}")
        search_text = " | ".join([normalize_text(p) for p in parts if p])
        final_nodes.append({
            "hs_code": hs,
            "title": title,
            "mo_ta_vn_samples": data["mo_ta_vn"][:2],
            "mo_ta_en_samples": data["mo_ta_en"][:2],
            "parents": sorted([p for p in data["parents"] if p]),
            "children": sorted(list(data["children"])),
            "other_samples": data["other_samples"][:3],
            "search_text": search_text
        })

    # grouped by top4
    grouped: Dict[str, Dict[str, Any]] = {}
    for rec in rows:
        code = rec.get("ma_hang_inferred")
        if not code:
            continue
        top4 = code[:4] if len(code) >= 4 else code
        if top4 not in grouped:
            grouped[top4] = {"root_code": top4, "root_rows": [], "children": []}
        def fallback_desc(rec, field):
            val = rec.get(field) or ""
            if val and str(val).strip() != "":
                return val
            v = normalize_text(rec.get("v_raw",""))
            v = re.sub(r'^\-+\s*', '', v).strip()
            return v
        mo_vn = fallback_desc(rec, "mo_ta_vn")
        mo_en = fallback_desc(rec, "mo_ta_en")
        if rec["v_level"] == 0 and len(code) == 4:
            grouped[top4]["root_rows"].append({
                "code": code,
                "desc": rec["v_raw"],
                "mo_ta_vn": mo_vn,
                "mo_ta_en": mo_en,
                "row_num": rec["row_num"]
            })
        else:
            grouped[top4]["children"].append({
                "code": code,
                "desc": rec["v_raw"],
                "mo_ta_vn": mo_vn,
                "mo_ta_en": mo_en,
                "level": rec["v_level"],
                "row_num": rec["row_num"]
            })

    # write outputs
    with open(output_rows, "w", encoding="utf-8") as f:
        json.dump(rows, f, ensure_ascii=False, indent=2)
    grouped_list = []
    for root in sorted(grouped.keys()):
        info = grouped[root]
        grouped_list.append({
            "root_code": root,
            "root_desc": info["root_rows"][0]["mo_ta_vn"] if info["root_rows"] and info["root_rows"][0].get("mo_ta_vn") else 
                         (info["children"][0]["mo_ta_vn"] if info["children"] and info["children"][0].get("mo_ta_vn") else ""),
            "root_rows": info["root_rows"],
            "children": info["children"]
        })
    with open(output_grouped, "w", encoding="utf-8") as f:
        json.dump(grouped_list, f, ensure_ascii=False, indent=2)
    with open(output_tree, "w", encoding="utf-8") as f:
        json.dump(final_nodes, f, ensure_ascii=False, indent=2)

    print(f"\nWrote {len(rows)} rows -> {output_rows}")
    print(f"Wrote {len(grouped_list)} groups -> {output_grouped}")
    print(f"Wrote {len(final_nodes)} unique nodes -> {output_tree}")

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", "-i", default="data/hscode.xlsx")
    parser.add_argument("--header-rows", "-hr", type=int, default=2)
    parser.add_argument("--output_rows", "-or", default="data/hscode_by_colletter.json")
    parser.add_argument("--output_grouped", "-og", default="data/hscode_grouped.json")
    parser.add_argument("--output_tree", "-ot", default="data/hscode_grouped_tree.json")
    parser.add_argument("--preview-only", action="store_true")
    args = parser.parse_args()

    process(
        input_path=args.input,
        header_rows=args.header_rows,
        output_rows=args.output_rows,
        output_grouped=args.output_grouped,
        output_tree=args.output_tree,
        preview_only=args.preview_only,
    )

if __name__ == "__main__":
    main()
