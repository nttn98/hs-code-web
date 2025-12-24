# excel_to_json.py
import json
import re
from pathlib import Path

from openpyxl import load_workbook


def depth_from_desc(desc: str) -> int:
    """
    Tính mức độ thụt theo số lượng dấu '-' ở đầu:
    "- ABC"  -> 1
    "- - ABC" -> 2
    "- - - ABC" -> 3
    ...
    """
    if not isinstance(desc, str):
        return 0
    s = desc.lstrip()
    depth = 0
    i = 0
    while i < len(s) and s[i] == "-":
        depth += 1
        i += 1
        if i < len(s) and s[i] == " ":
            i += 1
    return depth


def normalize_name(desc: str) -> str:
    """
    Bỏ dấu '-' ở đầu và dấu ':' ở cuối.
    Ví dụ: "- Ngựa:" -> "Ngựa"
           "- - Loại khác" -> "Loại khác"
    """
    if not isinstance(desc, str):
        return ""
    s = desc.lstrip()
    while s.startswith("-"):
        s = s[1:].lstrip()
    if s.endswith(":"):
        s = s[:-1].rstrip()
    return s


def excel_to_tree(path: str):
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    roots = []          # danh sách các Chương
    current_root = None # root hiện tại (Chương)
    current_child1 = None  # hàng in đậm (mã 4 số) hiện tại
    stack = {}          # lưu node theo depth (1,2,3,4,...)

    for r in range(1, ws.max_row + 1):
        cell_code = ws.cell(row=r, column=1)
        cell_desc = ws.cell(row=r, column=2)

        code_raw = cell_code.value
        desc_raw = cell_desc.value

        # bỏ qua dòng trống
        if code_raw is None and desc_raw is None:
            continue

        desc_str = str(desc_raw) if desc_raw is not None else None

        # 1. Dòng có chữ "Chương" -> root
        if desc_str and "Chương" in desc_str:
            m = re.search(r"Chương\s+(\d+)", desc_str)
            chap_id = m.group(1) if m else desc_str.strip()

            current_root = {
                "id": str(chap_id),
                "name": desc_str.strip(),  # "Chương 1"
                "description": "",         # sẽ điền ở dòng tiếp theo
                "leaf": []                 # danh sách child1 (hàng in đậm)
            }
            roots.append(current_root)
            current_child1 = None
            stack = {}
            continue

        # 2. Dòng mô tả chương (ngay sau dòng "Chương X", không có mã)
        if current_root and not current_root["description"] and desc_str and not code_raw:
            current_root["description"] = desc_str.strip()
            continue

        # nếu chưa có chương nào thì bỏ
        if not current_root:
            continue

        # 3. Hàng in đậm là child1 (mã 4 số, ví dụ 0101, 0102,...)
        is_bold = bool(cell_desc.font.bold or cell_code.font.bold)
        if (
            is_bold
            and code_raw not in (None, "", "Mã hàng")
            and desc_str
            and "Chương" not in desc_str
        ):
            node = {
                "id": str(code_raw).strip(),   # ví dụ "0101"
                "name": desc_str.strip(),      # "Ngựa, lừa, la sống."
                "group1": []                   # danh sách các nhóm "- Ngựa:", "- Lừa:", ...
            }
            current_root["leaf"].append(node)
            current_child1 = node
            stack = {}  # reset stack mỗi khi sang child1 mới
            continue

        # 4. Các dòng có '-', '--', '---'... là child2, child3, child4...
        depth = depth_from_desc(desc_str)
        if depth > 0 and current_child1:
            node = {
                "id": str(code_raw).strip() if code_raw not in (None, "") else None,
                "name": normalize_name(desc_str),
            }

            if depth == 1:
                # child2 trực tiếp của child1 => bỏ vào group1
                parent = current_child1
                parent.setdefault("group1", [])
                parent["group1"].append(node)
            else:
                # depth >= 2
                # parent là node ở depth-1
                parent = stack.get(depth - 1) or current_child1
                parent.setdefault("children", [])
                parent["children"].append(node)

            # lưu node này ở depth hiện tại
            stack[depth] = node
            continue

        # còn lại bỏ qua (dòng trống/không hợp lệ)

    # Hàm đệ quy bỏ các mảng children/group1 rỗng cho gọn JSON
    def prune(node: dict):
        if "group1" in node:
            for child in node["group1"]:
                prune(child)
            if not node["group1"]:
                node.pop("group1")
        if "children" in node:
            for child in node["children"]:
                prune(child)
            if not node["children"]:
                node.pop("children")

    for root in roots:
        for leaf in root["leaf"]:
            prune(leaf)

    return roots


if __name__ == "__main__":
    # Đổi đường dẫn này thành file Excel của bạn
    input_path = "./data/data.xlsx"
    output_path = "./data/hscode_tree.json"

    tree = excel_to_tree(input_path)

    # Nếu bạn muốn giống ví dụ 1 root duy nhất:
    # data = {"root": tree[0]}  # chương đầu tiên
    # Còn nếu muốn tất cả chương:
    data = {"roots": tree}

    Path(output_path).write_text(
        json.dumps(data, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )
    print(f"Đã ghi JSON ra: {output_path}")
