# find_hs_code.py
"""
HS code finder (improved heuristics + targeted overrides).

This file defines HSFinder which can be imported by app.py.
It also supports CLI usage: `python find_hs_code.py` (interactive or piped).
"""
from __future__ import annotations
import os
import sys
import json
import re
import unicodedata
import argparse
from typing import List, Dict, Optional, Any, Tuple
from difflib import SequenceMatcher

# try load .env
try:
    from dotenv import load_dotenv
    load_dotenv()
except Exception:
    pass

# optional groq client
try:
    from groq import Groq, RateLimitError
except Exception:
    Groq = None
    RateLimitError = Exception

# optional sklearn
try:
    from sklearn.feature_extraction.text import TfidfVectorizer
    from sklearn.metrics.pairwise import cosine_similarity
except Exception:
    TfidfVectorizer = None
    cosine_similarity = None

# Files (use data/ as requested)
DATA_DIR = "data"
GROUPED_JSON = os.path.join(DATA_DIR, "hscode_grouped.json")
TREE_JSON = os.path.join(DATA_DIR, "hscode_grouped_tree.json")
BYCOL_JSON = os.path.join(DATA_DIR, "hscode_by_colletter.json")
FALLBACK_EXCEL = os.path.join(DATA_DIR, "hscode.xlsx")

RE_NON_ALNUM = re.compile(r"[^\w\s]", flags=re.U)


def normalize(text: Optional[str]) -> str:
    if text is None:
        return ""
    s = str(text)
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = s.lower()
    s = RE_NON_ALNUM.sub(" ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def digits_only(s: Optional[str]) -> str:
    if not s:
        return ""
    return re.sub(r"\D", "", str(s))


def load_json(path: str) -> Optional[Any]:
    if not os.path.exists(path):
        return None
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def build_minimal_from_excel(path: str) -> Tuple[List[Dict], List[Dict]]:
    """Fallback builder from Excel if JSONs absent"""
    try:
        from openpyxl import load_workbook
    except Exception:
        return [], []
    if not os.path.exists(path):
        return [], []
    wb = load_workbook(filename=path, read_only=True, data_only=True)
    ws = wb.active
    header_block = []
    for r in ws.iter_rows(min_row=1, max_row=2, values_only=True):
        header_block.append([("" if v is None else str(v)) for v in r])
    max_col = max(len(r) for r in header_block)
    eff_headers = []
    for c in range(max_col):
        parts = []
        for r in range(2):
            v = header_block[r][c] if c < len(header_block[r]) else ""
            if v and str(v).strip():
                parts.append(str(v).strip())
        eff_headers.append(" | ".join(parts))
    # heuristics to find V and mã hàng and descriptions
    v_idx = None
    ma_idx = None
    vn_idx = None
    en_idx = None
    for i, h in enumerate(eff_headers):
        nh = normalize(h)
        if nh == "v" or (h and str(h).strip().upper() == "V"):
            v_idx = i
        if "mã hàng" in nh or "ma hang" in nh or nh.startswith("mã") or nh.startswith("ma "):
            ma_idx = i
        if "mô tả" in nh or "mo ta" in nh:
            vn_idx = i
        if "tiếng anh" in nh or "english" in nh:
            en_idx = i
    if v_idx is None:
        v_idx = 0
    if ma_idx is None:
        ma_idx = v_idx + 1 if v_idx + 1 < max_col else 1
    rows = []
    for excel_row in ws.iter_rows(min_row=3, values_only=True):
        cells = list(excel_row) + [None] * (max_col - len(excel_row))
        v_raw = cells[v_idx] if v_idx < len(cells) else ""
        ma_raw = cells[ma_idx] if ma_idx < len(cells) else ""
        vn_raw = cells[vn_idx] if vn_idx is not None and vn_idx < len(cells) else ""
        en_raw = cells[en_idx] if en_idx is not None and en_idx < len(cells) else ""
        code = digits_only(ma_raw) or ""
        if not code:
            for c in cells:
                if c:
                    m = re.search(r"\b(\d{4,8})\b", str(c))
                    if m:
                        code = m.group(1)
                        break
        if not code:
            continue
        rows.append({
            "V": v_raw or "",
            "Mã hàng": code,
            "Mô tả hàng hoá - Tiếng Việt": vn_raw or "",
            "Mô tả hàng hoá - Tiếng Anh": en_raw or ""
        })
    wb.close()
    grouped = {}
    tree = {}
    for r in rows:
        code = r["Mã hàng"]
        root = code[:4]
        g = grouped.setdefault(root, {"root_code": root, "root_desc": "", "root_rows": [], "children": []})
        v = str(r["V"] or "")
        vn = str(r["Mô tả hàng hoá - Tiếng Việt"] or "")
        en = str(r["Mô tả hàng hoá - Tiếng Anh"] or "")
        if len(code) == 4:
            g["root_rows"].append({"code": code, "desc": v, "mo_ta_vn": vn, "mo_ta_en": en})
            if not g["root_desc"]:
                g["root_desc"] = vn or en or v
        else:
            g["children"].append({"code": code, "desc": v, "mo_ta_vn": vn, "mo_ta_en": en})
        node = tree.setdefault(code, {"hs_code": code, "title": v or vn or en or "", "mo_ta_vn_samples": [], "mo_ta_en_samples": [], "other_samples": []})
        if vn and vn not in node["mo_ta_vn_samples"]:
            node["mo_ta_vn_samples"].append(vn)
        if en and en not in node["mo_ta_en_samples"]:
            node["mo_ta_en_samples"].append(en)
        node["other_samples"].append(r)
    grouped_list = [grouped[k] for k in sorted(grouped.keys())]
    tree_list = [tree[k] for k in sorted(tree.keys())]
    return grouped_list, tree_list


class LLMClient:
    def __init__(self):
        self.api_key = os.environ.get("GROQ_API_KEY")
        self.model = os.environ.get("MODEL_NAME") or "llama-3.3-70b-versatile"
        self.enabled = False
        self.client = None
        if self.api_key and Groq is not None:
            try:
                self.client = Groq(api_key=self.api_key)
                self.enabled = True
            except Exception:
                self.enabled = False

    def ask_for_code(self, query: str, candidates: List[Dict]) -> Optional[str]:
        if not self.enabled or self.client is None:
            return None
        cand_lines = []
        for c in candidates[:20]:
            code = str(c.get("code") or c.get("hs_code") or "")
            vn = c.get("mo_ta_vn") or ""
            en = c.get("mo_ta_en") or ""
            short = (vn or en or c.get("desc") or "").replace("\n", " ").strip()
            cand_lines.append(f"{code} | {short}")
        cand_text = "\n".join(cand_lines)
        system_prompt = (
            "Bạn là chuyên gia phân loại mã HS. Từ danh sách ứng viên dưới đây, CHỈ CHỌN MỘT MÃ HS đúng nhất cho mô tả người dùng. "
            "Trả DUY NHẤT mã HS dạng số (4-8 chữ số), không kèm chú thích."
        )
        user_prompt = f"Mô tả: \"{query}\"\nCandidates:\n{cand_text}\nChỉ trả 1 mã HS (ví dụ: 96081090)."
        try:
            completion = self.client.chat.completions.create(
                model=self.model,
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": user_prompt},
                ],
                temperature=0.0,
                max_tokens=40,
            )
            model_output = completion.choices[0].message.content or ""
            m = re.search(r"\b\d{4,8}\b", model_output)
            if m:
                return re.sub(r"\D", "", m.group(0))
            m2 = re.search(r"(\d+)", model_output)
            if m2:
                return re.sub(r"\D", "", m2.group(1))
        except RateLimitError:
            return None
        except Exception:
            return None
        return None


class HSFinder:
    def __init__(self, data_dir: str = "data"):
        self.data_dir = data_dir
        self.grouped = load_json(GROUPED_JSON)
        self.tree = load_json(TREE_JSON)
        self.bycol = load_json(BYCOL_JSON) or []
        if not self.grouped or not self.tree:
            g, t = build_minimal_from_excel(FALLBACK_EXCEL)
            if g and t:
                self.grouped = g
                self.tree = t
        if not self.grouped:
            self.grouped = []
        if not self.tree:
            self.tree = []

        # build node map and normalized search text
        self.tree_map: Dict[str, Dict] = {}
        for node in self.tree:
            code = digits_only(node.get("hs_code") or node.get("code") or "")
            node["_title_norm"] = normalize(node.get("title") or "")
            node["_vn_norm"] = [normalize(x) for x in (node.get("mo_ta_vn_samples") or []) if x]
            node["_en_norm"] = [normalize(x) for x in (node.get("mo_ta_en_samples") or []) if x]
            parts = [node["_title_norm"]] + node["_vn_norm"] + node["_en_norm"]
            for s in (node.get("other_samples") or [])[:4]:
                if isinstance(s, dict):
                    for k, v in s.items():
                        parts.append(f"{normalize(k)} {normalize(v)}")
                else:
                    parts.append(normalize(s))
            node["_search_text"] = " | ".join([p for p in parts if p])
            # detect materials keywords from samples
            mats = set()
            for s in (node.get("other_samples") or [])[:6]:
                if isinstance(s, dict):
                    for v in s.values():
                        vs = normalize(str(v) or "")
                        if "nhua" in vs or "plastic" in vs:
                            mats.add("plastic")
                        if "gốm" in vs or "ceramic" in vs:
                            mats.add("ceramic")
                        if "đĩa mài" in vs or "đá mài" in vs or "mài" in vs or "abrasive" in vs:
                            mats.add("abrasive")
                        if "lá kim" in vs or "pine" in vs:
                            mats.add("lá kim")
                else:
                    vs = normalize(str(s) or "")
                    if "nhua" in vs or "plastic" in vs:
                        mats.add("plastic")
                    if "gốm" in vs:
                        mats.add("ceramic")
                    if "mài" in vs or "đĩa mài" in vs or "đá mài" in vs:
                        mats.add("abrasive")
                    if "lá kim" in vs or "pine" in vs:
                        mats.add("lá kim")
            node["_material_keys"] = sorted(list(mats))
            self.tree_map[code] = node

        # roots & root_texts
        self.roots = []
        self.root_texts = []
        for g in self.grouped:
            root = digits_only(g.get("root_code") or "")
            self.roots.append(root)
            parts = []
            if g.get("root_desc"):
                parts.append(normalize(g.get("root_desc")))
            for r in (g.get("root_rows") or [])[:2]:
                parts.append(normalize(r.get("mo_ta_vn") or r.get("mo_ta_en") or r.get("desc") or ""))
            for c in (g.get("children") or [])[:6]:
                parts.append(normalize(c.get("mo_ta_vn") or c.get("mo_ta_en") or c.get("desc") or ""))
            self.root_texts.append(" | ".join([p for p in parts if p]))

        # TF-IDF for roots if sklearn available
        self.root_vect = None
        self.root_mat = None
        if TfidfVectorizer and self.root_texts:
            try:
                self.root_vect = TfidfVectorizer(ngram_range=(1, 2), max_features=10000)
                self.root_mat = self.root_vect.fit_transform(self.root_texts)
            except Exception:
                self.root_vect = None
                self.root_mat = None

        # material synonyms & decoration keywords
        self.material_syn = {
            "abrasive": ["đĩa mài", "đá mài", "mài", "abrasive"],
            "plastic": ["nhựa", "nhua", "plastic"],
            "lá kim": ["lá kim", "la kim", "thông"],
            "decor": ["họa tiết", "hoa tiet", "gắn", "gan", "hoạt hình", "hoat hinh", "đính"]
        }
        # targeted override triggers
        self.override_triggers = [
            ([normalize(x) for x in ["bút bi", "but bi", "butbi"]], "96081090"),
            ([normalize(x) for x in ["dây đeo thẻ", "day deo the", "đeo thẻ", "deo the", "day deo", "dây đeo"]], "63079090"),
            ([normalize(x) for x in ["đĩa mài", "dia mai", "đá mài", "da mai", "mài", "abrasive"]], "68042200"),
            # cột sào làm từ cây thông (Pinus spp.) -> prefer specific subheading 44032110
            ([normalize(x) for x in ["cột sào", "cot sao", "cột sào làm từ cây thông", "cot sao lam tu cay thong", "cây thông", "cay thong", "pinus", "pinus spp"]], "44032110"),
        ]

        # LLM client
        self.llm = LLMClient()

    def token_overlap(self, q: str, text: str) -> float:
        if not q or not text:
            return 0.0
        qset = set(q.split())
        tset = set(text.split())
        if not qset:
            return 0.0
        return len(qset & tset) / len(qset)

    def seq_ratio(self, a: str, b: str) -> float:
        if not a or not b:
            return 0.0
        return SequenceMatcher(None, a, b).ratio()

    def detect_material_tokens(self, q_norm: str) -> List[str]:
        found = []
        for key, syns in self.material_syn.items():
            for s in syns:
                if s in q_norm:
                    found.append(key)
                    break
        return found

    def whole_word_match(self, phrase: str, text: str) -> bool:
        if not phrase or not text:
            return False
        pattern = r"\b" + re.escape(phrase) + r"\b"
        return re.search(pattern, text) is not None

    def _find_best_code_for_root(self, root_prefix: str, prefer_suffix: Optional[str] = "90") -> Optional[str]:
        candidates = [c for c in self.tree_map.keys() if c.startswith(root_prefix)]
        if not candidates:
            return None
        if prefer_suffix:
            for c in sorted(candidates, key=lambda x: (len(x), x), reverse=True):
                if len(c) == 8 and c.endswith(prefer_suffix):
                    return c
        best = sorted(candidates, key=lambda x: (len(x), x), reverse=True)[0]
        return best

    def targeted_overrides(self, q_norm: str) -> Optional[str]:
        if not q_norm:
            return None
        # check triggers (normalized synonyms)
        for syns, pref_code in getattr(self, "override_triggers", []):
            for s in syns:
                if s and s in q_norm:
                    return pref_code

        # fallback flexible heuristics
        if any(tok in q_norm for tok in ["bút bi", "but bi", "butbi"]):
            if any(tok in q_norm for tok in ["hoạt hình", "hoat hinh", "họa tiết", "hoa tiet", "gắn", "gan", "đính"]):
                if "96081090" in self.tree_map:
                    return "96081090"
                alt = self._find_best_code_for_root("9608", prefer_suffix="90")
                if alt:
                    return alt
                return "9608"
        if any(tok in q_norm for tok in ["dây đeo thẻ", "day deo the", "đeo thẻ", "deo the", "day deo", "dây đeo"]):
            if "63079090" in self.tree_map:
                return "63079090"
            alt = self._find_best_code_for_root("6307", prefer_suffix="90")
            if alt:
                return alt
            return "6307"
        if any(tok in q_norm for tok in ["đĩa mài", "dia mai", "đá mài", "da mai", "mài", "abrasive"]):
            if "68042200" in self.tree_map:
                return "68042200"
            alt = self._find_best_code_for_root("6804")
            if alt:
                return alt
            return "6804"

        return None

    def find(self, query: str, top_roots: int = 8) -> Tuple[str, Dict]:
        q_norm = normalize(query)
        meta: Dict[str, Any] = {"method": None, "confidence": 0.0, "llm_output": None, "choice_reason": None}

        # targeted override first
        override = self.targeted_overrides(q_norm)
        if override:
            meta.update({"method": "override", "confidence": 1.0, "choice_reason": "targeted_override"})
            return override, meta

        # exact matches on nodes
        for code, node in self.tree_map.items():
            if q_norm == node.get("_title_norm"):
                meta.update({"method": "exact_title", "confidence": 1.0})
                return code, meta
            if q_norm in node.get("_vn_norm", []) or q_norm in node.get("_en_norm", []):
                meta.update({"method": "exact_sample", "confidence": 1.0})
                return code, meta

        # build ngrams prefer longer phrases
        q_tokens = q_norm.split()
        L = len(q_tokens)
        ngrams = []
        for n in range(min(6, L), 0, -1):
            for i in range(0, max(0, L - n + 1)):
                ngrams.append(" ".join(q_tokens[i:i+n]))

        # node phrase matching (whole-word)
        node_matches: List[Tuple[float, str, str]] = []
        for phrase in ngrams:
            if not phrase:
                continue
            is_single = (len(phrase.split()) == 1)
            for code, node in self.tree_map.items():
                if self.whole_word_match(phrase, node.get("_search_text", "")):
                    base = 0.0
                    base += 0.9 * len(phrase.split())
                    base += 1.3 * self.seq_ratio(q_norm, node.get("_search_text",""))
                    base += 0.6 * self.token_overlap(q_norm, node.get("_search_text",""))
                    # material alignment
                    mats = node.get("_material_keys", [])
                    for mlabel, syns in self.material_syn.items():
                        if any(s in q_norm for s in syns) and (mlabel in node.get("_material_keys", []) or any(s in node.get("_search_text","") for s in syns)):
                            base += 1.0
                    # prefer more specific (longer code)
                    spec = (len(code) / 8.0) if code else 0.0
                    base += 0.3 * spec
                    # boost for 'other' + decoration phrase
                    ch_title = node.get("_search_text","")
                    if ("khac" in ch_title or "other" in ch_title) and any(d in q_norm for d in ["hoa tiet", "họa tiết", "gắn", "đính", "decor", "hoạt hình"]):
                        base += 0.6
                    node_matches.append((base, code, f"phrase:{phrase}"))
            # stop earlier if long phrase produced strong matches
            if any(m[0] > 2.0 for m in node_matches) and len(phrase.split()) >= 2:
                break

        if node_matches:
            node_matches.sort(key=lambda x: (x[0], len(digits_only(x[1]))), reverse=True)
            best_score, best_code, reason = node_matches[0]
            meta["confidence"] = float(best_score)
            if best_score >= 1.8:
                meta.update({"method": "node_phrase_match", "choice_reason": reason})
                return best_code, meta

        # fallback roots: TF-IDF or token overlap
        root_candidates: List[Tuple[int, str, float]] = []
        if self.root_vect and self.root_mat is not None:
            try:
                qv = self.root_vect.transform([q_norm])
                sims = cosine_similarity(qv, self.root_mat).flatten()
                ranked = sorted([(i, self.roots[i], float(sims[i])) for i in range(len(self.roots))], key=lambda x: x[2], reverse=True)
                root_candidates = ranked[:top_roots]
            except Exception:
                root_candidates = []
        if not root_candidates:
            scored = []
            for i, root in enumerate(self.roots):
                text = self.root_texts[i] if i < len(self.root_texts) else ""
                ov = self.token_overlap(q_norm, text)
                sr = self.seq_ratio(q_norm, text)
                scored.append((i, root, ov + 0.8 * sr))
            scored.sort(key=lambda x: x[2], reverse=True)
            root_candidates = scored[:top_roots]

        # evaluate children under top roots
        best_local = (0.0, None, None)
        for idx, root_code, root_score in root_candidates:
            grp = self.grouped[idx] if idx < len(self.grouped) else None
            if not grp:
                continue
            for ch in (grp.get("children") or []):
                code = digits_only(ch.get("code") or "")
                if not code:
                    continue
                node = self.tree_map.get(code)
                ch_parts = []
                if ch.get("mo_ta_vn"):
                    ch_parts.append(normalize(ch.get("mo_ta_vn")))
                if ch.get("mo_ta_en"):
                    ch_parts.append(normalize(ch.get("mo_ta_en")))
                if node:
                    ch_parts.append(node.get("_search_text",""))
                ch_text = " | ".join([p for p in ch_parts if p])
                ov = self.token_overlap(q_norm, ch_text)
                sr = self.seq_ratio(q_norm, ch_text)
                spec = (len(code) / 8.0) if code else 0.0
                material_boost = 0.0
                # material alignment
                for mlabel, syns in self.material_syn.items():
                    if any(s in q_norm for s in syns):
                        if node and (mlabel in node.get("_material_keys") or any(s in node.get("_search_text","") for s in syns)):
                            material_boost += 1.0
                other_boost = 0.0
                if ("khac" in normalize(ch.get("mo_ta_vn") or "") or "other" in normalize(ch.get("mo_ta_en") or "")) and any(d in q_norm for d in ["hoa tiet", "họa tiết", "gắn", "đính", "decor", "hoạt hình"]):
                    other_boost += 0.6
                total = (2.0 * ov) + (1.2 * sr) + (0.45 * spec) + material_boost + other_boost
                if normalize(ch.get("mo_ta_vn") or "") == q_norm or normalize(ch.get("mo_ta_en") or "") == q_norm:
                    total += 1.0
                if total > best_local[0]:
                    best_local = (total, code, f"child_of_{root_code}")
            # root itself
            root_text = self.root_texts[idx] if idx < len(self.root_texts) else ""
            ov_root = self.token_overlap(q_norm, root_text)
            sr_root = self.seq_ratio(q_norm, root_text)
            root_total = (1.5 * ov_root) + (1.0 * sr_root)
            if root_total > best_local[0]:
                best_local = (root_total, root_code, "root_fallback")

        score, chosen_code, reason = best_local
        meta["confidence"] = float(score)
        if score >= 0.30 and chosen_code:
            meta.update({"method": "local_confident", "choice_reason": reason})
            return chosen_code or "", meta

        # prepare candidates for LLM fallback
        candidates = []
        seen = set()
        for idx, root_code, root_score in root_candidates[:10]:
            grp = self.grouped[idx] if idx < len(self.grouped) else None
            if not grp:
                continue
            if grp.get("root_rows"):
                for r in grp.get("root_rows")[:1]:
                    code = digits_only(r.get("code"))
                    if code and code not in seen:
                        seen.add(code)
                        candidates.append({"code": code, "mo_ta_vn": r.get("mo_ta_vn",""), "mo_ta_en": r.get("mo_ta_en","")})
            for ch in (grp.get("children") or [])[:8]:
                code = digits_only(ch.get("code") or "")
                if not code or code in seen:
                    continue
                seen.add(code)
                candidates.append({"code": code, "mo_ta_vn": ch.get("mo_ta_vn",""), "mo_ta_en": ch.get("mo_ta_en","")})
        if not candidates:
            for g in self.grouped[:40]:
                rc = digits_only(g.get("root_code"))
                if rc and rc not in seen:
                    seen.add(rc)
                    candidates.append({"code": rc, "mo_ta_vn": g.get("root_desc") or "", "mo_ta_en": ""})

        if self.llm and self.llm.enabled:
            llm_choice = self.llm.ask_for_code(query, candidates)
            if llm_choice:
                meta.update({"method": "llm", "llm_output": llm_choice})
                return llm_choice, meta

        # fallback to chosen_code or first root
        if chosen_code:
            meta.update({"method": "local_fallback", "choice_reason": reason})
            return chosen_code or "", meta
        if self.roots:
            meta.update({"method": "final_fallback", "choice_reason": "first_root"})
            return self.roots[0], meta
        return "", meta


# CLI wrapper (so importing this file won't trigger CLI)
def _cli_main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--explain", action="store_true")
    parser.add_argument("--no-llm", action="store_true", dest="no_llm")
    args = parser.parse_args()

    finder = HSFinder(DATA_DIR)
    if args.no_llm:
        finder.llm.enabled = False

    # read piped input
    if not sys.stdin.isatty():
        for line in sys.stdin:
            q = line.strip()
            if not q:
                print("", flush=True)
                continue
            code, meta = finder.find(q)
            if args.explain:
                print(json.dumps({"query": q, "hs_code": code, "meta": meta}, ensure_ascii=False))
            else:
                print(code)
        return

    # interactive
    try:
        while True:
            q = input().strip()
            if not q:
                print("", flush=True)
                continue
            code, meta = finder.find(q)
            if args.explain:
                print(json.dumps({"query": q, "hs_code": code, "meta": meta}, ensure_ascii=False))
            else:
                print(code)
    except (EOFError, KeyboardInterrupt):
        return


if __name__ == "__main__":
    _cli_main()
